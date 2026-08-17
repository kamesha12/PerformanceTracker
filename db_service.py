import pyodbc
import io
import logging
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from config import (
    SQLSERVER_HOST,
    SQLSERVER_PORT,
    SQLSERVER_NAME,
    SQLSERVER_USER,
    SQLSERVER_PASSWORD,
    SQLSERVER_DRIVER,
    SQLSERVER_TRUSTED_CONNECTION,
    build_connection_string
)

logger = logging.getLogger("app_logger")

EXPECTED_COLUMNS = [
    "S.No.",
    "Intern Name",
    "Branch",
    "Physical Visits",
    "Telecalling",
    "DSA/Connectors",
    "Promoters/Builders",
    "Weekly Visits",
    "Leads Achieved",
    "Marketing Activity",
    "Insight"
]

INITIAL_RECORDS = [
    (1, 'Devadharshini M', 'Nungambakkam, Chennai', 1, 28, 0, 0, 1, 3, 0, 'Needs focus on physical visits.'),
    (2, 'Kavya P', 'Nungambakkam, Chennai', 1, 28, 0, 0, 1, 2, 0, 'Consistent telecalling effort.'),
    (3, 'Swetha S', 'Nungambakkam, Chennai', 1, 28, 0, 0, 1, 2, 0, 'Good steady performance.'),
    (4, 'Sandhya S', 'Nungambakkam, Chennai', 1, 28, 0, 0, 1, 1, 0, 'Solid communication skills.'),
    (5, 'Aadhithyan SV', 'West Zone - Mumbai', 40, 20, 20, 20, 40, 45, 10, 'High performing lead converter.'),
    (6, 'Balaji N', 'West Zone - Mumbai', 35, 15, 15, 15, 35, 38, 8, 'Strong promoter network.'),
    (7, 'Prasanth R', 'South Zone - Bengaluru', 42, 25, 25, 20, 42, 42, 12, 'Excellent weekly visit record.'),
    (8, 'Rithik S', 'South Zone - Bengaluru', 30, 18, 12, 10, 30, 28, 5, 'Consistent lead pipeline.'),
    (9, 'Santhosh K', 'North Zone - Delhi', 25, 30, 10, 8, 25, 22, 6, 'Very strong telecalling stats.'),
    (10, 'Vishwa M', 'North Zone - Delhi', 20, 35, 8, 5, 20, 19, 4, 'High customer outreach.'),
    (11, 'Dinesh Kumar', 'South Zone - Hyderabad', 38, 22, 18, 14, 38, 40, 9, 'Top performer in region.'),
    (12, 'Gokul R', 'South Zone - Hyderabad', 28, 20, 14, 11, 28, 26, 7, 'Steady weekly improvement.')
]

_cached_conn_str = None

def get_db_connection(database=None):
    """
    Attempts to establish a pyodbc connection to Microsoft SQL Server
    with fallback driver, server, and authentication options.
    """
    global _cached_conn_str
    target_db = database if database else SQLSERVER_NAME

    if _cached_conn_str and database is None:
        try:
            return pyodbc.connect(_cached_conn_str, timeout=3)
        except Exception:
            _cached_conn_str = None

    available_drivers = pyodbc.drivers()

    drivers_to_try = []
    if SQLSERVER_DRIVER in available_drivers:
        drivers_to_try.append(SQLSERVER_DRIVER)
    for d in ['ODBC Driver 18 for SQL Server', 'ODBC Driver 17 for SQL Server', 'SQL Server']:
        if d in available_drivers and d not in drivers_to_try:
            drivers_to_try.append(d)

    servers_to_try = [
        SQLSERVER_HOST,
        r"(localdb)\MSSQLLocalDB",
        "127.0.0.1",
        "localhost",
        r".\SQLEXPRESS",
        r"localhost\SQLEXPRESS"
    ]

    last_error = None
    for drv in drivers_to_try:
        for srv in servers_to_try:
            if SQLSERVER_TRUSTED_CONNECTION.lower() not in ("yes", "true", "1") and SQLSERVER_USER:
                conn_str = f"DRIVER={{{drv}}};SERVER={srv};DATABASE={target_db};UID={SQLSERVER_USER};PWD={SQLSERVER_PASSWORD};TrustServerCertificate=yes;"
                try:
                    conn = pyodbc.connect(conn_str, timeout=1)
                    if database is None:
                        _cached_conn_str = conn_str
                    return conn
                except Exception as e:
                    last_error = e

            conn_str_trusted = f"DRIVER={{{drv}}};SERVER={srv};DATABASE={target_db};Trusted_Connection=yes;TrustServerCertificate=yes;"
            try:
                conn = pyodbc.connect(conn_str_trusted, timeout=1)
                if database is None:
                    _cached_conn_str = conn_str_trusted
                return conn
            except Exception as e:
                last_error = e

    logger.error(f"Failed to connect to SQL Server database ({target_db}): {last_error}")
    raise RuntimeError(f"Failed to connect to Microsoft SQL Server: {last_error}")

def init_db():
    """
    Initializes SQL Server database and table, inserting initial records if empty.
    Non-blocking on startup.
    """
    try:
        # Step 1: Ensure company_db Database Exists
        try:
            conn = get_db_connection(database="master")
            conn.autocommit = True
            cursor = conn.cursor()
            cursor.execute("SELECT database_id FROM sys.databases WHERE name = ?", (SQLSERVER_NAME,))
            if not cursor.fetchone():
                cursor.execute(f"CREATE DATABASE [{SQLSERVER_NAME}]")
                logger.info(f"Created SQL Server database '{SQLSERVER_NAME}'.")
            cursor.close()
            conn.close()
        except Exception as err:
            logger.warning(f"Database creation check against master skipped/failed: {err}")

        # Step 2: Ensure dbo.intern_performance Table Exists
        conn = get_db_connection(database=SQLSERVER_NAME)
        conn.autocommit = True
        cursor = conn.cursor()

        create_table_sql = """
        IF OBJECT_ID('dbo.intern_performance', 'U') IS NULL
        BEGIN
            CREATE TABLE dbo.intern_performance (
                s_no INT IDENTITY(1,1) PRIMARY KEY,
                intern_name NVARCHAR(255) NOT NULL,
                branch NVARCHAR(255),
                physical_visits INT DEFAULT 0,
                telecalling INT DEFAULT 0,
                dsa_connectors INT DEFAULT 0,
                promoters_builders INT DEFAULT 0,
                weekly_visits INT DEFAULT 0,
                leads_achieved INT DEFAULT 0,
                marketing_activity INT DEFAULT 0,
                insight NVARCHAR(MAX)
            );
        END
        """
        cursor.execute(create_table_sql)

        # Step 3: Check if Table is Empty and Seed Data
        cursor.execute("SELECT COUNT(*) FROM dbo.intern_performance")
        count = cursor.fetchone()[0]

        if count == 0:
            logger.info("Seeding initial 12 intern performance records into SQL Server...")
            cursor.execute("SET IDENTITY_INSERT dbo.intern_performance ON;")
            
            insert_sql = """
            INSERT INTO dbo.intern_performance 
            (s_no, intern_name, branch, physical_visits, telecalling, dsa_connectors, promoters_builders, weekly_visits, leads_achieved, marketing_activity, insight)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?);
            """
            cursor.executemany(insert_sql, INITIAL_RECORDS)
            cursor.execute("SET IDENTITY_INSERT dbo.intern_performance OFF;")
            logger.info("Successfully seeded SQL Server table with initial records.")

        cursor.close()
        conn.close()
        logger.info("SQL Server Database Initialization Complete.")
    except Exception as e:
        logger.warning(f"SQL Server Server check on startup: {e}")

def get_all_records(search_query="", sort_by="", sort_order="asc"):
    """
    Fetches all intern records from Microsoft SQL Server database.
    """
    try:
        conn = get_db_connection()
        cursor = conn.cursor()

        sql = """
        SELECT 
            s_no AS [S.No.],
            intern_name AS [Intern Name],
            branch AS [Branch],
            physical_visits AS [Physical Visits],
            telecalling AS [Telecalling],
            dsa_connectors AS [DSA/Connectors],
            promoters_builders AS [Promoters/Builders],
            weekly_visits AS [Weekly Visits],
            leads_achieved AS [Leads Achieved],
            marketing_activity AS [Marketing Activity],
            insight AS [Insight]
        FROM dbo.intern_performance
        """

        params = []
        if search_query:
            sql += " WHERE LOWER(intern_name) LIKE ? OR LOWER(branch) LIKE ?"
            term = f"%{search_query.lower()}%"
            params.extend([term, term])

        # Dynamic Sorting
        sort_map = {
            "name": "[Intern Name]",
            "physical": "[Physical Visits]",
            "telecalling": "[Telecalling]",
            "leads": "[Leads Achieved]",
            "marketing": "[Marketing Activity]"
        }

        order_col = sort_map.get(sort_by, "[S.No.]")
        direction = "ASC" if sort_order.lower() == "asc" else "DESC"
        sql += f" ORDER BY {order_col} {direction}"

        cursor.execute(sql, params)
        columns = [column[0] for column in cursor.description]
        rows = cursor.fetchall()

        records = []
        for row in rows:
            record_dict = dict(zip(columns, row))
            records.append(record_dict)

        cursor.close()
        conn.close()

        # Compute Executive Summary KPIs
        total_interns = len(records)
        total_physical_visits = sum(r.get("Physical Visits", 0) or 0 for r in records)
        total_telecalling = sum(r.get("Telecalling", 0) or 0 for r in records)
        total_weekly_visits = sum(r.get("Weekly Visits", 0) or 0 for r in records)
        total_leads_achieved = sum(r.get("Leads Achieved", 0) or 0 for r in records)
        total_marketing_activities = sum(r.get("Marketing Activity", 0) or 0 for r in records)
        total_dsa_connectors = sum(r.get("DSA/Connectors", 0) or 0 for r in records)
        total_promoters_builders = sum(r.get("Promoters/Builders", 0) or 0 for r in records)

        summary = {
            "total_interns": total_interns,
            "total_physical_visits": total_physical_visits,
            "total_telecalling": total_telecalling,
            "total_weekly_visits": total_weekly_visits,
            "total_leads_achieved": total_leads_achieved,
            "total_marketing_activities": total_marketing_activities,
            "total_dsa_connectors": total_dsa_connectors,
            "total_promoters_builders": total_promoters_builders
        }

        # Donut Chart Data
        branch_leads = {}
        for r in records:
            b = r.get("Branch", "Unknown") or "Unknown"
            l = r.get("Leads Achieved", 0) or 0
            branch_leads[b] = branch_leads.get(b, 0) + l

        donut_labels = list(branch_leads.keys())
        donut_values = list(branch_leads.values())
        donut_sum = sum(donut_values) if donut_values else 1
        donut_data = [
            {
                "labels": label,
                "values": val,
                "percentage": round((val / (donut_sum if donut_sum > 0 else 1)) * 100, 1)
            }
            for label, val in zip(donut_labels, donut_values)
        ]

        # Bar Chart Telecalling
        top_tele = sorted(records, key=lambda x: x.get("Telecalling", 0) or 0, reverse=True)[:10]
        bar_telecalling = [
            {"names": r.get("Intern Name", "N/A"), "values": r.get("Telecalling", 0) or 0}
            for r in top_tele
        ]

        # Line Chart Physical Visits
        line_physical = [
            {"names": r.get("Intern Name", "N/A"), "values": r.get("Physical Visits", 0) or 0}
            for r in records[:10]
        ]

        # Bar Chart Marketing
        top_marketing = sorted(records, key=lambda x: x.get("Marketing Activity", 0) or 0, reverse=True)[:10]
        bar_marketing = [
            {"names": r.get("Intern Name", "N/A"), "values": r.get("Marketing Activity", 0) or 0}
            for r in top_marketing
        ]

        charts = {
            "donut_leads": donut_data,
            "bar_telecalling": bar_telecalling,
            "line_physical_visits": line_physical,
            "bar_marketing": bar_marketing
        }

        return {
            "summary": summary,
            "charts": charts,
            "records": records,
            "columns": EXPECTED_COLUMNS
        }
    except Exception as e:
        logger.error(f"Cannot fetch records from SQL Server: {e}")
        return {
            "summary": {},
            "charts": {},
            "records": [],
            "columns": EXPECTED_COLUMNS
        }

def db_add_record(payload):
    """
    Inserts a new intern record into Microsoft SQL Server database.
    """
    conn = get_db_connection()
    conn.autocommit = True
    cursor = conn.cursor()

    def parse_int(val):
        try:
            return int(val)
        except (ValueError, TypeError):
            return 0

    sql = """
    INSERT INTO dbo.intern_performance 
    (intern_name, branch, physical_visits, telecalling, dsa_connectors, promoters_builders, weekly_visits, leads_achieved, marketing_activity, insight)
    OUTPUT INSERTED.s_no
    VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?);
    """

    params = (
        str(payload.get("Intern Name", "")).strip(),
        str(payload.get("Branch", "")).strip(),
        parse_int(payload.get("Physical Visits", 0)),
        parse_int(payload.get("Telecalling", 0)),
        parse_int(payload.get("DSA/Connectors", 0)),
        parse_int(payload.get("Promoters/Builders", 0)),
        parse_int(payload.get("Weekly Visits", 0)),
        parse_int(payload.get("Leads Achieved", 0)),
        parse_int(payload.get("Marketing Activity", 0)),
        str(payload.get("Insight", "")).strip()
    )

    cursor.execute(sql, params)
    new_id = cursor.fetchone()[0]
    cursor.close()
    conn.close()

    result_record = dict(payload)
    result_record["S.No."] = new_id

    logger.info(f"Inserted record #{new_id} into SQL Server successfully.")
    return result_record

def db_update_record(s_no, payload):
    """
    Updates an existing intern record in Microsoft SQL Server database.
    """
    conn = get_db_connection()
    conn.autocommit = True
    cursor = conn.cursor()

    cursor.execute("SELECT s_no FROM dbo.intern_performance WHERE s_no = ?", (s_no,))
    if not cursor.fetchone():
        cursor.close()
        conn.close()
        raise ValueError(f"Record with S.No. #{s_no} not found.")

    def parse_int(val):
        try:
            return int(val)
        except (ValueError, TypeError):
            return 0

    sql = """
    UPDATE dbo.intern_performance
    SET 
        intern_name = ?,
        branch = ?,
        physical_visits = ?,
        telecalling = ?,
        dsa_connectors = ?,
        promoters_builders = ?,
        weekly_visits = ?,
        leads_achieved = ?,
        marketing_activity = ?,
        insight = ?
    WHERE s_no = ?;
    """

    params = (
        str(payload.get("Intern Name", "")).strip(),
        str(payload.get("Branch", "")).strip(),
        parse_int(payload.get("Physical Visits", 0)),
        parse_int(payload.get("Telecalling", 0)),
        parse_int(payload.get("DSA/Connectors", 0)),
        parse_int(payload.get("Promoters/Builders", 0)),
        parse_int(payload.get("Weekly Visits", 0)),
        parse_int(payload.get("Leads Achieved", 0)),
        parse_int(payload.get("Marketing Activity", 0)),
        str(payload.get("Insight", "")).strip(),
        s_no
    )

    cursor.execute(sql, params)
    cursor.close()
    conn.close()

    logger.info(f"Updated record #{s_no} in SQL Server successfully.")
    return True

def db_delete_record(s_no):
    """
    Deletes an intern record from Microsoft SQL Server database.
    """
    conn = get_db_connection()
    conn.autocommit = True
    cursor = conn.cursor()

    cursor.execute("SELECT s_no FROM dbo.intern_performance WHERE s_no = ?", (s_no,))
    if not cursor.fetchone():
        cursor.close()
        conn.close()
        raise ValueError(f"Record with S.No. #{s_no} not found.")

    cursor.execute("DELETE FROM dbo.intern_performance WHERE s_no = ?", (s_no,))
    cursor.close()
    conn.close()

    logger.info(f"Deleted record #{s_no} from SQL Server successfully.")
    return True

def db_bulk_delete_records(s_nos):
    """
    Bulk deletes multiple intern records from Microsoft SQL Server.
    """
    if not s_nos:
        return 0
    conn = get_db_connection()
    conn.autocommit = True
    cursor = conn.cursor()
    placeholders = ",".join(["?"] * len(s_nos))
    sql = f"DELETE FROM dbo.intern_performance WHERE s_no IN ({placeholders});"
    cursor.execute(sql, s_nos)
    count = cursor.rowcount
    cursor.close()
    conn.close()
    logger.info(f"Bulk deleted {count} records from SQL Server.")
    return count

def db_bulk_update_records(s_nos, update_data):
    """
    Bulk updates specific fields for multiple intern records in Microsoft SQL Server.
    """
    if not s_nos or not update_data:
        return 0
    conn = get_db_connection()
    conn.autocommit = True
    cursor = conn.cursor()

    col_name_map = {
        "Branch": "branch",
        "Physical Visits": "physical_visits",
        "Telecalling": "telecalling",
        "DSA/Connectors": "dsa_connectors",
        "Promoters/Builders": "promoters_builders",
        "Weekly Visits": "weekly_visits",
        "Leads Achieved": "leads_achieved",
        "Marketing Activity": "marketing_activity",
        "Insight": "insight"
    }

    set_clauses = []
    params = []

    for key, val in update_data.items():
        db_col = col_name_map.get(key)
        if db_col:
            set_clauses.append(f"{db_col} = ?")
            if "Visits" in key or "Calling" in key or "Connectors" in key or "Builders" in key or "Leads" in key or "Activity" in key:
                try:
                    params.append(int(val))
                except (ValueError, TypeError):
                    params.append(0)
            else:
                params.append(str(val).strip())

    if not set_clauses:
        cursor.close()
        conn.close()
        return 0

    placeholders = ",".join(["?"] * len(s_nos))
    sql = f"UPDATE dbo.intern_performance SET {', '.join(set_clauses)} WHERE s_no IN ({placeholders});"
    params.extend(s_nos)

    cursor.execute(sql, params)
    count = cursor.rowcount
    cursor.close()
    conn.close()
    logger.info(f"Bulk updated {count} records in SQL Server.")
    return count

def get_excel_bytes():
    """
    Generates an in-memory Excel spreadsheet (.xlsx BytesIO buffer)
    directly from Microsoft SQL Server database records.
    """
    data = get_all_records()
    records = data.get("records", [])

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Performance Tracker"

    # Styling
    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    header_font = Font(name="Segoe UI", size=11, bold=True, color="FFFFFF")
    data_font = Font(name="Segoe UI", size=10)
    thin_border = Border(
        left=Side(style='thin', color='D9D9D9'),
        right=Side(style='thin', color='D9D9D9'),
        top=Side(style='thin', color='D9D9D9'),
        bottom=Side(style='thin', color='D9D9D9')
    )

    # Write Header Row
    ws.append(EXPECTED_COLUMNS)
    for col_num in range(1, len(EXPECTED_COLUMNS) + 1):
        cell = ws.cell(row=1, column=col_num)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # Write Data Rows
    for row_idx, r in enumerate(records, start=2):
        row_data = [r.get(col, "") for col in EXPECTED_COLUMNS]
        ws.append(row_data)

        for col_num in range(1, len(EXPECTED_COLUMNS) + 1):
            cell = ws.cell(row=row_idx, column=col_num)
            cell.font = data_font
            cell.border = thin_border
            col_name = EXPECTED_COLUMNS[col_num - 1]
            if col_name in ["S.No.", "Physical Visits", "Telecalling", "DSA/Connectors", "Promoters/Builders", "Weekly Visits", "Leads Achieved", "Marketing Activity"]:
                cell.alignment = Alignment(horizontal="center", vertical="center")
            else:
                cell.alignment = Alignment(horizontal="left", vertical="center")

    # Auto-fit Column Widths
    for col in ws.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        col_letter = get_column_letter(col[0].column)
        ws.column_dimensions[col_letter].width = max(max_len + 3, 12)

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer
